from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import config

def create_excel_report(hosts, scan_info, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"
    
    # Стили
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    color_fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') 
                  for k, v in config.COLORS.items()}
    
    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.append(["Общая информация"])
    ws.append(["Время запуска сканирования:", scan_info.get('start_time', 'N/A')])
    ws.append(["Хост-источник:", scan_info.get('source_host', 'N/A')])
    ws.append(["Выполненная команда:", scan_info.get('command', 'N/A')])
    ws.append(["Обработано хостов:", f"{scan_info.get('hosts_up', 0)} (из {scan_info.get('total_ips', 0)})"])
    ws.append([])
    
    # Заголовки хостов
    headers = ["Хосты:"] + [host['ip'] for host in hosts]
    ws.append(headers)
    
    # Hostnames
    hostnames = ["hostname:"] + [host.get('hostname', '') for host in hosts]
    ws.append(hostnames)
    ws.append([])  # Пустая строка перед портами
    
    # Собираем все уникальные порты
    all_ports = sorted(set(
        port 
        for host in hosts 
        for port in host['ports'].keys()
    ), key=lambda x: (int(x[0]) if x[0].isdigit() else 0, x[1], x[2]))
    
    # Заголовки для портов и сервисов
    ws.append(["Порты", "Протокол", "Сервис"] + [""] * len(hosts))
    
    # Добавляем порты
    for port_info in all_ports:
        port, protocol, service = port_info
        row = [port, protocol, service]
        for host in hosts:
            state = host['ports'].get(port_info, 'undefined')
            row.append(state)
        ws.append(row)
    
    # Применяем форматирование
    # 1. Объединяем только нужные ячейки
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) + 1)  # "Общая информация"
    
    # 2. Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=9 + len(all_ports), max_col=len(headers) + 2):
        for cell in row:
            cell.font = header_font if cell.row <= 9 else Font()
            cell.border = thin_border
            if cell.row == 1:
                cell.alignment = alignment_center
    
    # 3. Добавляем цвета для состояний портов
    for row_idx in range(10, 10 + len(all_ports)):
        # Заголовки портов
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).font = header_font
        
        # Состояния портов
        for col_idx in range(4, len(headers) + 3):
            state_cell = ws.cell(row=row_idx, column=col_idx)
            state = state_cell.value.lower() if state_cell.value else ''
            
            if 'open' in state:
                state_cell.fill = color_fills['open']
            elif 'closed' in state:
                state_cell.fill = color_fills['closed']
            elif 'filtered' in state:
                state_cell.fill = color_fills['filtered']
            elif 'undefined' in state:
                state_cell.fill = color_fills['undefined']
            else:
                state_cell.fill = color_fills['default']
    
    # 4. Настраиваем ширину колонок
    col_widths = {}
    for col_idx in range(1, len(headers) + 3):
        max_length = 0
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_widths[col_idx] = max(10, min(30, max_length + 2))
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 5. Центрирование текста в заголовках
    for row in [7, 8, 9]:  # Строки с заголовками
        for col in range(1, len(headers) + 3):
            ws.cell(row=row, column=col).alignment = alignment_center
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"Основной отчёт сохранён как: {output_file}")
    
    # Создаем транспонированную версию
    create_transposed_sheet(output_file, hosts, scan_info, all_ports)

def create_transposed_sheet(output_file, hosts, scan_info, all_ports):
    # Загружаем существующую книгу
    wb = load_workbook(output_file)
    
    # Создаем новый лист для транспонированных данных
    ws = wb.create_sheet("Transposed View")
    
    # Стили
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    color_fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') 
                  for k, v in config.COLORS.items()}
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.append(["Транспонированное представление"])
    ws.append(["Время запуска сканирования:", scan_info.get('start_time', 'N/A')])
    ws.append(["Хост-источник:", scan_info.get('source_host', 'N/A')])
    ws.append(["Выполненная команда:", scan_info.get('command', 'N/A')])
    ws.append(["Обработано хостов:", f"{scan_info.get('hosts_up', 0)} (из {scan_info.get('total_ips', 0)})"])
    ws.append([])
    
    # Заголовки портов
    port_headers = ["IP", "Hostname"]
    for port_info in all_ports:
        port, protocol, service = port_info
        port_headers.append(f"{port}/{protocol}")
        port_headers.append(service)
    
    ws.append(port_headers)
    
    # Добавляем хосты
    for host in hosts:
        row = [host['ip'], host.get('hostname', '')]
        for port_info in all_ports:
            state = host['ports'].get(port_info, 'undefined')
            row.append(state)
            row.append("")  # Пустая ячейка для сервиса
        ws.append(row)
    
    # Применяем форматирование
    # 1. Объединяем только нужные ячейки
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(port_headers))
    
    # 2. Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=7 + len(hosts), max_col=len(port_headers)):
        for cell in row:
            cell.font = header_font if cell.row <= 7 else Font()
            cell.border = thin_border
            if cell.row == 1:
                cell.alignment = alignment_center
    
    # 3. Добавляем цвета для состояний портов
    for row_idx in range(8, 8 + len(hosts)):
        # Заголовки хостов
        for col_idx in range(1, 3):
            ws.cell(row=row_idx, column=col_idx).font = header_font
        
        # Состояния портов (каждый 3-й столбец начиная с 3)
        for col_idx in range(3, len(port_headers) + 1, 2):
            state_cell = ws.cell(row=row_idx, column=col_idx)
            state = state_cell.value.lower() if state_cell.value else ''
            
            if 'open' in state:
                state_cell.fill = color_fills['open']
            elif 'closed' in state:
                state_cell.fill = color_fills['closed']
            elif 'filtered' in state:
                state_cell.fill = color_fills['filtered']
            elif 'undefined' in state:
                state_cell.fill = color_fills['undefined']
            else:
                state_cell.fill = color_fills['default']
    
    # 4. Настраиваем ширину колонок
    col_widths = {}
    for col_idx in range(1, len(port_headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_widths[col_idx] = max(10, min(30, max_length + 2))
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"Транспонированное представление добавлено в: {output_file}")
