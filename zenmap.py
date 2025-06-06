import subprocess
import socket
import re
import os
import sys
import argparse
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config import COLORS, DEFAULT_OUTPUT_DIR, OUTPUT_SUFFIX, SCAN_PROFILES, NMAP_PATH,SCAN_DIR

def parse_nmap_txt(file_path):
    hosts = []
    scan_info = {
        'start_time': None,
        'command': None,
        'source_host': None,
        'total_ips': 0,
        'hosts_up': 0
    }
    
    current_host = None
    host_pattern = re.compile(r'Nmap scan report for (?:([\w\-. ]+)\s)?\(?([\d.]+)\)?')
    ports_section = False
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            # Парсинг общей информации
            if not scan_info['start_time'] and line.startswith('Starting Nmap'):
                scan_info['start_time'] = line.split(' at ')[-1]
            
            elif 'Nmap done:' in line:
                match = re.search(r'Nmap done: (\d+) IP addresses \((\d+) hosts? up\)', line)
                if match:
                    scan_info['total_ips'] = int(match.group(1))
                    scan_info['hosts_up'] = int(match.group(2))
            
            # Парсинг хостов
            host_match = host_pattern.match(line)
            if host_match:
                if current_host:
                    hosts.append(current_host)
                
                hostname, ip = host_match.groups()
                current_host = {
                    'ip': ip,
                    'hostname': hostname.strip() if hostname else None,
                    'ports': {}
                }
                ports_section = False
                continue
            
            if current_host is None:
                continue
            
            # Начало секции с портами
            if line.startswith('PORT') and 'STATE' in line and 'SERVICE' in line:
                ports_section = True
                continue
            
            # Конец секции с портами
            if ports_section and (not line or line.startswith('Nmap scan')):
                ports_section = False
            
            # Парсинг портов
            if ports_section and line:
                # Пропускаем информационные строки
                if line.startswith('Not shown:') or line.startswith('All ') or 'filtered' in line:
                    continue
                
                # Исправленный парсинг строк портов
                parts = re.split(r'\s+', line, maxsplit=2)
                if len(parts) < 2:
                    continue
                
                port_service = parts[0].strip()
                state = parts[1].strip().lower()
                service = parts[2].strip().split()[0] if len(parts) > 2 else 'unknown'
                
                if '/' not in port_service:
                    continue
                    
                port, protocol = port_service.split('/', 1)
                
                # Сохраняем как отдельные компоненты
                port_key = (port, protocol, service)
                current_host['ports'][port_key] = state

    # Добавляем последний хост
    if current_host:
        hosts.append(current_host)
    
    return hosts, scan_info

def create_excel_report(hosts, scan_info, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"
    
    # Стили
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    color_fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') for k, v in COLORS.items()}
    
    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.append(["Общая информация"])
    ws.append(["Время запуска сканирования:", scan_info.get('start_time', 'N/A')])
    ws.append(["Хост-источник:", scan_info.get('source_host', 'N/A')])
    ws.append(["Выполненная команда:", scan_info.get('command', 'N/A')])
    ws.append(["Обработано хостов:", f"{scan_info.get('hosts_up', 0)} (из {scan_info.get('total_ips', 0)})"])
    ws.append([])
    
    # Заголовки хостов
    headers = ["Хосты:"] + [host['ip'] for host in hosts]
    ws.append(headers)
    
    # Hostnames
    hostnames = ["hostname:"] + [host.get('hostname', '') for host in hosts]
    ws.append(hostnames)
    ws.append([])  # Пустая строка перед портами
    
    # Собираем все уникальные порты
    all_ports = sorted(set(
        port 
        for host in hosts 
        for port in host['ports'].keys()
    ), key=lambda x: (int(x[0]) if x[0].isdigit() else 0, x[1], x[2]))

    
    # Заголовки для портов и сервисов
    ws.append(["Порты", "Протокол", "Сервис"] + [""] * len(hosts))
    
    # Добавляем порты
    for port_info in all_ports:
        port, protocol, service = port_info
        row = [port, protocol, service]
        for host in hosts:
            state = host['ports'].get(port_info, 'undefined')
            row.append(state)
        ws.append(row)
    
    # Применяем форматирование
    # 1. Объединяем только нужные ячейки
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) + 1)  # "Общая информация"
    
    # 2. Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=9 + len(all_ports), max_col=len(headers) + 2):
        for cell in row:
            cell.font = header_font if cell.row <= 9 else Font()
            cell.border = thin_border
            if cell.row == 1:
                cell.alignment = alignment_center
    
    # 3. Добавляем цвета для состояний портов
    for row_idx in range(10, 10 + len(all_ports)):
        # Заголовки портов
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).font = header_font
        
        # Состояния портов
        for col_idx in range(4, len(headers) + 3):
            state_cell = ws.cell(row=row_idx, column=col_idx)
            state = state_cell.value.lower() if state_cell.value else ''
            
            if 'open' in state:
                state_cell.fill = color_fills['open']
            elif 'closed' in state:
                state_cell.fill = color_fills['closed']
            elif 'filtered' in state:
                state_cell.fill = color_fills['filtered']
            elif 'undefined' in state:
                state_cell.fill = color_fills['undefined']
            else:
                state_cell.fill = color_fills['default']
    
    # 4. Настраиваем ширину колонок
    col_widths = {}
    for col_idx in range(1, len(headers) + 3):
        max_length = 0
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_widths[col_idx] = max(10, min(30, max_length + 2))
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 5. Центрирование текста в заголовках
    for row in [7, 8, 9]:  # Строки с заголовками
        for col in range(1, len(headers) + 3):
            ws.cell(row=row, column=col).alignment = alignment_center
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"Основной отчёт сохранён как: {output_file}")
    
    # Создаем транспонированную версию
    create_transposed_sheet(output_file, hosts, scan_info, all_ports)

def create_transposed_sheet(output_file, hosts, scan_info, all_ports):
    # Загружаем существующую книгу
    wb = load_workbook(output_file)
    
    # Создаем новый лист для транспонированных данных
    ws = wb.create_sheet("Transposed View")
    
    # Стили
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    color_fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') for k, v in COLORS.items()}
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.append(["Транспонированное представление"])
    ws.append(["Время запуска сканирования:", scan_info.get('start_time', 'N/A')])
    ws.append(["Хост-источник:", scan_info.get('source_host', 'N/A')])
    ws.append(["Выполненная команда:", scan_info.get('command', 'N/A')])
    ws.append(["Обработано хостов:", f"{scan_info.get('hosts_up', 0)} (из {scan_info.get('total_ips', 0)})"])
    ws.append([])
    
    # Заголовки портов
    port_headers = ["IP", "Hostname"]
    for port_info in all_ports:
        port, protocol, service = port_info
        port_headers.append(f"{port}/{protocol}")
        port_headers.append(service)
    
    ws.append(port_headers)
    
    # Добавляем хосты
    for host in hosts:
        row = [host['ip'], host.get('hostname', '')]
        for port_info in all_ports:
            state = host['ports'].get(port_info, 'undefined')
            row.append(state)
            row.append("")  # Пустая ячейка для сервиса
        ws.append(row)
    
    # Применяем форматирование
    # 1. Объединяем только нужные ячейки
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(port_headers))
    
    # 2. Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=7 + len(hosts), max_col=len(port_headers)):
        for cell in row:
            cell.font = header_font if cell.row <= 7 else Font()
            cell.border = thin_border
            if cell.row == 1:
                cell.alignment = alignment_center
    
    # 3. Добавляем цвета для состояний портов
    for row_idx in range(8, 8 + len(hosts)):
        # Заголовки хостов
        for col_idx in range(1, 3):
            ws.cell(row=row_idx, column=col_idx).font = header_font
        
        # Состояния портов (каждый 3-й столбец начиная с 3)
        for col_idx in range(3, len(port_headers) + 1, 2):
            state_cell = ws.cell(row=row_idx, column=col_idx)
            state = state_cell.value.lower() if state_cell.value else ''
            
            if 'open' in state:
                state_cell.fill = color_fills['open']
            elif 'closed' in state:
                state_cell.fill = color_fills['closed']
            elif 'filtered' in state:
                state_cell.fill = color_fills['filtered']
            elif 'undefined' in state:
                state_cell.fill = color_fills['undefined']
            else:
                state_cell.fill = color_fills['default']
    
    # 4. Настраиваем ширину колонок
    col_widths = {}
    for col_idx in range(1, len(port_headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_widths[col_idx] = max(10, min(30, max_length + 2))
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"Транспонированное представление добавлено в: {output_file}")

def get_local_ip():
    """Получаем локальный IP-адрес"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "N/A"

def run_nmap_scan(target, profile, output_file):
    """Запускает Nmap сканирование с выбранным профилем"""
    profiles = {
        "ping":                         [NMAP_PATH, "-sn", target, "-oN", output_file],
        "intense":                      [NMAP_PATH, "-T4", "-A", "-v", target, "-oN", output_file],
        "Intense plus UDP":             [NMAP_PATH, "-sS", "-sU", "-T4", "-A", "-v", target, "-oN", output_file],
        "Intense scan, all TCP ports":  [NMAP_PATH, "-p", "1-65535", "-T4", "-A", "-v", target, "-oN", output_file],
        "Intense scan, no ping":        [NMAP_PATH, "-sn", target, "-oN", output_file],
        "Quick scan":                   [NMAP_PATH, "-sn", target, "-oN", output_file],
        "Quick scan plus":              [NMAP_PATH, "-sn", target, "-oN", output_file],
        "Quick traceroute":             [NMAP_PATH, "-sn", target, "-oN", output_file],
        "Regular scan":                 [NMAP_PATH, "-sn", target, "-oN", output_file],
        "Slow comprehensive scan":      [NMAP_PATH, "-sn", target, "-oN", output_file]
    }
    
    if profile not in profiles:
        raise ValueError(f"Неизвестный профиль: {profile}")
    
    command = profiles[profile]
    print(f"Выполняем команду: {' '.join(command)}")
    
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        return True
    except subprocess.CalledProcessError as e:
        print(f"Ошибка при выполнении сканирования (код {e.returncode}):")
        print(e.stderr)
        return False
    except Exception as e:
        print(f"Неожиданная ошибка: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description='Обработчик результатов сканирования Nmap')
    parser.add_argument('input_file', nargs='?', help='Путь к файлу с результатами сканирования Nmap')
    parser.add_argument('--output-dir', help='Папка для сохранения отчета')
    parser.add_argument('--target', help='Цель для сканирования (например, 192.168.1.0/24)')
    parser.add_argument('--profile', choices=list(SCAN_PROFILES.keys()), 
                        help=f'Профиль сканирования: {", ".join(SCAN_PROFILES.keys())}')
    
    args = parser.parse_args()
    
    # Если указаны target и profile, запускаем сканирование
    if args.target and args.profile:
        # Создаем имя файла для сканирования
        os.makedirs(SCAN_DIR, exist_ok=True)
        scan_file = os.path.join(SCAN_DIR, f"nmap_scan_{args.profile}.txt")
        
        if run_nmap_scan(args.target, args.profile, scan_file):
            print(f"Сканирование завершено. Результаты сохранены в: {scan_file}")
            input_file = scan_file
            scan_info = {
                'command': f"nmap {args.profile} {args.target}",
                'source_host': get_local_ip()
            }
        else:
            print("Ошибка при выполнении сканирования")
            sys.exit(3)
    
    # Определяем папку для вывода
    output_dir = args.output_dir or DEFAULT_OUTPUT_DIR
    
    # Проверка существования файла
    if not os.path.exists(input_file):
        print(f"Ошибка: файл '{input_file}' не найден")
        sys.exit(1)
    
    # Определение пути для выходного файла
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file_name = f"{base_name}{OUTPUT_SUFFIX}"
    
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_file_name)
    else:
        output_path = os.path.join(os.path.dirname(__file__), output_file_name)
    
    try:
        # Если scan_info не был создан при сканировании, парсим файл
        if not scan_info:
            hosts, scan_info = parse_nmap_txt(input_file)
        else:
            hosts, scan_info_from_file = parse_nmap_txt(input_file)
            # Объединяем информацию из сканирования и из файла
            scan_info.update({
                'start_time': scan_info_from_file.get('start_time'),
                'total_ips': scan_info_from_file.get('total_ips'),
                'hosts_up': scan_info_from_file.get('hosts_up')
            })
        
        if not hosts:
            print("Предупреждение: не найдено данных для обработки")
            wb = Workbook()
            wb.save(output_path)
            print(f"Создан пустой отчет: {output_path}")
        else:
            create_excel_report(hosts, scan_info, output_path)
            
    except Exception as e:
        print(f"Критическая ошибка при обработке файла: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(2)

if __name__ == "__main__":
    main()
